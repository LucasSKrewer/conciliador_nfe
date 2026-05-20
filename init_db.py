"""
init_db.py — Conciliador NF-e

Cria o banco SQLite e importa as 2 planilhas:
  * SEFAZ (.xlsx)    — notas lançadas CONTRA a empresa (export do FSist
                       no formato "FSist-NFe-Recebidas-*.xlsx" ou similar)
  * Sistema (.csv)   — notas lançadas no seu sistema interno (ERP),
                       em CSV ponto-e-vírgula com a coluna "Chave"

A importação é IDEMPOTENTE: pode ser rodada várias vezes sem perder as
marcações manuais de "cartão" nem observações. O matching é feito pela
chave NF-e de 44 dígitos.

Uso:
    python init_db.py
    python init_db.py --sefaz "caminho\\FSist...xlsx" --sistema "caminho\\notas.csv"
"""

import argparse
import csv
import os
import re
import sqlite3
import sys
from datetime import datetime, date

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except (AttributeError, OSError):
    pass

import openpyxl

DB_PATH       = "conciliador.db"
SCHEMA_PATH   = "schema.sql"
ARQUIVO_SEFAZ_PADRAO   = None   # detectado automaticamente na pasta atual
ARQUIVO_SISTEMA_PADRAO = "Notas de Entrada.csv"

# ── helpers ──────────────────────────────────────────────────────────────────

def _s(v):
    """Converte valor para string limpa, ou None."""
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None

def _chave(v):
    """Normaliza chave NFe: só dígitos, 44 caracteres."""
    if v is None:
        return None
    s = re.sub(r"\D", "", str(v))
    return s if len(s) == 44 else None

def _cnpj(v):
    """Normaliza CNPJ: só dígitos."""
    if v is None:
        return None
    s = re.sub(r"\D", "", str(v))
    return s if s else None

def _data_iso(v):
    """Aceita datetime/date/str e devolve 'YYYY-MM-DD' ou None."""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    # tenta dd/mm/yyyy ou dd/mm/yy
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{2,4})", s)
    if m:
        d, mo, y = m.groups()
        if len(y) == 2:
            y = ("20" + y) if int(y) < 70 else ("19" + y)
        return f"{y}-{int(mo):02d}-{int(d):02d}"
    # tenta yyyy-mm-dd
    m = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})", s)
    if m:
        y, mo, d = m.groups()
        return f"{y}-{int(mo):02d}-{int(d):02d}"
    return None

def _parse_emitida_por(s):
    """Recebe 'XX.XXX.XXX/XXXX-XX - RAZÃO SOCIAL' (CNPJ + traço com espaços + nome)
    e devolve (cnpj_digits, razao). Separador é ' - ' (espaço-traço-espaço) pra
    não confundir com o '-XX' do CNPJ."""
    if not s:
        return None, None
    s = str(s).strip()
    m = re.match(r"^\s*([\d./\-]+)\s+-\s+(.+)$", s)
    if m:
        cnpj_dig = re.sub(r"\D", "", m.group(1))
        razao = m.group(2).strip()
        return (cnpj_dig if len(cnpj_dig) in (11, 14) else None), razao
    return None, s

def _chave_nfse_url(url):
    """Extrai os 50 dígitos finais do URL DANFSe (nfse.gov.br/.../DANFSe/<chave>)."""
    if not url:
        return None
    digits = re.sub(r"\D", "", str(url))
    return digits[-50:] if len(digits) >= 50 else None

def _doc_no_nome(s):
    """Extrai sequência de 11-14 dígitos no final da razão social (CPF/CNPJ embutido)."""
    if not s:
        return None
    m = re.search(r"(\d{11,14})\s*$", str(s).strip())
    return m.group(1) if m else None

def _is_cancelada(status):
    """True se o status/situação indica nota cancelada (ignora maiúsculas/acentos)."""
    if not status:
        return False
    return "cancela" in str(status).lower()

def _norm_emitente(s):
    """Normaliza razão social para comparação: uppercase, sem acento, só A-Z0-9."""
    if not s:
        return ""
    import unicodedata
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode()
    return re.sub(r"[^A-Z0-9]", "", s.upper())

def _num_br(v):
    """Converte número em formato BR ('1.234,56') ou normal para float."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    # remove separador de milhar (.) e troca decimal , por .
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None

# ── banco ────────────────────────────────────────────────────────────────────

def criar_banco(conn):
    with open(SCHEMA_PATH, "r", encoding="utf-8") as f:
        conn.executescript(f.read())
    _aplicar_migrations(conn)
    conn.commit()

def _aplicar_migrations(conn):
    """Migrations idempotentes para bancos criados antes de novas colunas."""
    cols_nota = {row[1] for row in conn.execute("PRAGMA table_info(nota_consolidada)")}
    if "usuario_lancamento" not in cols_nota:
        conn.execute("ALTER TABLE nota_consolidada ADD COLUMN usuario_lancamento TEXT")

    cols_imp = {row[1] for row in conn.execute("PRAGMA table_info(importacao)")}
    for col, decl in [
        ("arquivo_cte",       "TEXT"),
        ("arquivo_nfse",      "TEXT"),
        ("ctes_sefaz",        "INTEGER NOT NULL DEFAULT 0"),
        ("ctes_sistema",      "INTEGER NOT NULL DEFAULT 0"),
        ("ctes_total",        "INTEGER NOT NULL DEFAULT 0"),
        ("nfses_prefeitura",  "INTEGER NOT NULL DEFAULT 0"),
        ("nfses_sistema",     "INTEGER NOT NULL DEFAULT 0"),
        ("nfses_total",       "INTEGER NOT NULL DEFAULT 0"),
    ]:
        if col not in cols_imp:
            conn.execute(f"ALTER TABLE importacao ADD COLUMN {col} {decl}")

def consolidar_nfs_com_nfes(conn):
    """Para cada NFS sintética em nota_consolidada (chave LIKE 'NFS-%'),
    tenta achar uma NFe (chave de 44 dígitos) com:
      1) mesmo número + razão social prefixo 8 chars
      2) ou mesmo número + valor exato (cobre razão diferente entre SEFAZ
         e ERP, ex: nome fantasia vs razão social)
    Se achar, transfere em_sistema/usuario/cartão/obs pra NFe e DELETA a NFS.
    """
    sinteticas = conn.execute("""
        SELECT chave, numero, emitente, valor, em_sistema, marcacao_cartao,
               observacao, usuario_lancamento
        FROM nota_consolidada WHERE chave LIKE 'NFS-%'
    """).fetchall()
    consolidadas = 0
    for s in sinteticas:
        chave_s, numero_s, emit_s, valor_s, em_sis, cartao, obs, usu = s
        num_dig = re.sub(r"[^0-9]", "", numero_s or "")
        if not num_dig:
            continue
        norm_s = _norm_emitente(emit_s or "")
        prefix = norm_s[:8] if len(norm_s) >= 6 else None
        candidatos = conn.execute("""
            SELECT chave, emitente, valor FROM nota_consolidada
            WHERE numero = ? AND length(chave) = 44
        """, (num_dig,)).fetchall()
        alvo = None
        if prefix:
            for c in candidatos:
                if _norm_emitente(c[1] or "").startswith(prefix):
                    alvo = c[0]; break
        if not alvo and valor_s is not None:
            for c in candidatos:
                if c[2] is not None and abs(float(valor_s) - float(c[2])) < 0.01:
                    alvo = c[0]; break
        if not alvo:
            continue
        conn.execute("""
            UPDATE nota_consolidada SET
                em_sistema = CASE WHEN ?=1 OR em_sistema=1 THEN 1 ELSE em_sistema END,
                marcacao_cartao = CASE WHEN ?=1 THEN 1 ELSE marcacao_cartao END,
                observacao = COALESCE(observacao, ?),
                usuario_lancamento = COALESCE(usuario_lancamento, ?),
                ultima_importacao = datetime('now','localtime')
            WHERE chave = ?
        """, (em_sis or 0, cartao or 0, obs, usu, alvo))
        conn.execute("DELETE FROM nota_consolidada WHERE chave=?", (chave_s,))
        consolidadas += 1
    conn.commit()
    return consolidadas

def migrar_nfs_sinteticas(conn):
    """Para cada NFS sintética em nota_consolidada, tenta achar NFSe
    correspondente em nfse_consolidada por CPF/CNPJ embutido ou razão
    social normalizada. Se achar, marca em_sistema=1 + usuario na NFSe
    e remove a NFS de nota_consolidada."""
    sinteticas = conn.execute("""
        SELECT chave, emitente, observacao, marcacao_cartao, usuario_lancamento
        FROM nota_consolidada WHERE chave LIKE 'NFS-%'
    """).fetchall()
    nfses = conn.execute("""
        SELECT chave, cnpj_emitente, emitente FROM nfse_consolidada WHERE em_prefeitura=1
    """).fetchall()
    idx_doc = {}
    idx_pref = {}
    for r in nfses:
        norm = _norm_emitente(r[2] or "")
        doc_embutido = _doc_no_nome(r[2] or "")
        if r[1]: idx_doc[r[1]] = r[0]
        if doc_embutido: idx_doc[doc_embutido] = r[0]
        if len(norm) >= 6:
            idx_pref.setdefault(norm[:8], []).append((r[0], norm))
    movidas = 0
    for s in sinteticas:
        chave_s, emit_s, obs_s, cartao_s, usu_s = s
        doc = _doc_no_nome(emit_s or "")
        norm = _norm_emitente(emit_s or "")
        alvo = None
        if doc and doc in idx_doc:
            alvo = idx_doc[doc]
        elif norm and len(norm) >= 6:
            for k, lst in idx_pref.items():
                if norm.startswith(k):
                    alvo = lst[0][0]; break
        if not alvo:
            continue
        conn.execute("""
            UPDATE nfse_consolidada SET
                em_sistema = 1,
                marcacao_cartao = CASE WHEN ?=1 THEN 1 ELSE marcacao_cartao END,
                observacao = COALESCE(observacao, ?),
                usuario_lancamento = COALESCE(usuario_lancamento, ?),
                ultima_importacao = datetime('now','localtime')
            WHERE chave = ?
        """, (cartao_s or 0, obs_s, usu_s, alvo))
        conn.execute("DELETE FROM nota_consolidada WHERE chave=?", (chave_s,))
        movidas += 1
    conn.commit()
    return movidas

# ── importadores ─────────────────────────────────────────────────────────────

def detectar_arquivo_sefaz(pasta="."):
    """Acha o arquivo FSist-NFe-Recebidas-*.xlsx mais recente."""
    candidatos = []
    for nome in os.listdir(pasta):
        if nome.lower().startswith("fsist-nfe-recebidas") and nome.lower().endswith(".xlsx"):
            candidatos.append(os.path.join(pasta, nome))
    if not candidatos:
        return None
    return max(candidatos, key=os.path.getmtime)

def detectar_arquivo_cte(pasta="."):
    """Acha o arquivo FSist-CTe-*.xlsx mais recente."""
    candidatos = []
    for nome in os.listdir(pasta):
        low = nome.lower()
        if low.startswith("fsist-cte") and low.endswith(".xlsx"):
            candidatos.append(os.path.join(pasta, nome))
    if not candidatos:
        return None
    return max(candidatos, key=os.path.getmtime)

def importar_sefaz(conn, caminho):
    """Lê o xlsx da SEFAZ (FSist) e marca em_sefaz=1 nas notas encontradas.

    Colunas esperadas (cabeçalho na linha 1):
      0  Emissão
      1  Emissão Data/Hora
      2  Chave
      3  Mês/Ano
      4  Número
      5  Série
      6  Tipo
      7  Valor
      8  Status
      9  Manifestação
      10 Emitente CNPJ
      11 Emitente
      ...
    """
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb.active
    inseridos = atualizados = ignorados = canceladas = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        chave = _chave(row[2])
        if not chave:
            ignorados += 1
            continue
        # col 8 = Status — pula canceladas e remove do banco se existir
        if _is_cancelada(row[8] if len(row) > 8 else None):
            conn.execute("DELETE FROM nota_consolidada WHERE chave=?", (chave,))
            canceladas += 1
            continue
        data_emissao = _data_iso(row[0])
        numero       = _s(row[4])
        serie        = _s(row[5])
        valor        = _num_br(row[7])
        cnpj_emit    = _cnpj(row[10])
        emitente     = _s(row[11])

        ja_existe = conn.execute(
            "SELECT 1 FROM nota_consolidada WHERE chave=?", (chave,)
        ).fetchone()

        conn.execute("""
            INSERT INTO nota_consolidada
                (chave, numero, serie, cnpj_emitente, emitente, valor,
                 data_emissao, em_sefaz, ultima_importacao)
            VALUES (?, ?, ?, ?, ?, ?, ?, 1, datetime('now','localtime'))
            ON CONFLICT(chave) DO UPDATE SET
                numero         = COALESCE(excluded.numero, nota_consolidada.numero),
                serie          = COALESCE(excluded.serie, nota_consolidada.serie),
                cnpj_emitente  = COALESCE(excluded.cnpj_emitente, nota_consolidada.cnpj_emitente),
                emitente       = COALESCE(excluded.emitente, nota_consolidada.emitente),
                valor          = COALESCE(excluded.valor, nota_consolidada.valor),
                data_emissao   = COALESCE(excluded.data_emissao, nota_consolidada.data_emissao),
                em_sefaz       = 1,
                ultima_importacao = datetime('now','localtime')
        """, (chave, numero, serie, cnpj_emit, emitente, valor, data_emissao))

        if ja_existe:
            atualizados += 1
        else:
            inseridos += 1

    wb.close()
    return {"inseridos": inseridos, "atualizados": atualizados,
            "ignorados": ignorados, "canceladas": canceladas}

def importar_cte(conn, caminho):
    """Lê o xlsx de CT-e (FSist-CTe) e marca em_sefaz=1 em cte_consolidada.

    Colunas esperadas (cabeçalho linha 1):
      0  Chave
      1  Emissão
      4  Número
      5  Série
      7  Modal
      8  Tipo Serviço
      9  Valor (do frete)
     12  Valor da Carga
     13  Emitente CNPJ      (transportadora)
     14  Emitente
     16  Emitente UF
     25  Remetente CNPJ/CPF (quem despachou)
     26  Remetente
     38  NFe Chaves (com vírgula)
    """
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb.active
    inseridos = atualizados = ignorados = canceladas = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        chave = _chave(row[0])
        if not chave:
            ignorados += 1
            continue
        # col 10 = Status — pula canceladas e remove do banco se existir
        if _is_cancelada(row[10] if len(row) > 10 else None):
            conn.execute("DELETE FROM cte_consolidada WHERE chave=?", (chave,))
            canceladas += 1
            continue
        data_emissao  = _data_iso(row[1])
        numero        = _s(row[4])
        serie         = _s(row[5])
        modal         = _s(row[7])
        tipo_servico  = _s(row[8])
        valor         = _num_br(row[9])
        valor_carga   = _num_br(row[12])
        cnpj_emit     = _cnpj(row[13])
        emitente      = _s(row[14])
        emitente_uf   = _s(row[16])
        cnpj_rem      = _cnpj(row[25])
        remetente     = _s(row[26])
        nfe_chaves    = _s(row[38])

        ja_existe = conn.execute(
            "SELECT 1 FROM cte_consolidada WHERE chave=?", (chave,)
        ).fetchone()

        conn.execute("""
            INSERT INTO cte_consolidada
                (chave, numero, serie, cnpj_emitente, emitente, emitente_uf,
                 modal, tipo_servico, valor, valor_carga,
                 cnpj_remetente, remetente, nfe_chaves,
                 data_emissao, em_sefaz, ultima_importacao)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, datetime('now','localtime'))
            ON CONFLICT(chave) DO UPDATE SET
                numero         = COALESCE(excluded.numero, cte_consolidada.numero),
                serie          = COALESCE(excluded.serie, cte_consolidada.serie),
                cnpj_emitente  = COALESCE(excluded.cnpj_emitente, cte_consolidada.cnpj_emitente),
                emitente       = COALESCE(excluded.emitente, cte_consolidada.emitente),
                emitente_uf    = COALESCE(excluded.emitente_uf, cte_consolidada.emitente_uf),
                modal          = COALESCE(excluded.modal, cte_consolidada.modal),
                tipo_servico   = COALESCE(excluded.tipo_servico, cte_consolidada.tipo_servico),
                valor          = COALESCE(excluded.valor, cte_consolidada.valor),
                valor_carga    = COALESCE(excluded.valor_carga, cte_consolidada.valor_carga),
                cnpj_remetente = COALESCE(excluded.cnpj_remetente, cte_consolidada.cnpj_remetente),
                remetente      = COALESCE(excluded.remetente, cte_consolidada.remetente),
                nfe_chaves     = COALESCE(excluded.nfe_chaves, cte_consolidada.nfe_chaves),
                data_emissao   = COALESCE(excluded.data_emissao, cte_consolidada.data_emissao),
                em_sefaz       = 1,
                ultima_importacao = datetime('now','localtime')
        """, (chave, numero, serie, cnpj_emit, emitente, emitente_uf,
              modal, tipo_servico, valor, valor_carga,
              cnpj_rem, remetente, nfe_chaves, data_emissao))

        if ja_existe:
            atualizados += 1
        else:
            inseridos += 1

    wb.close()
    return {"inseridos": inseridos, "atualizados": atualizados,
            "ignorados": ignorados, "canceladas": canceladas}

def detectar_arquivo_nfse(pasta="."):
    """Acha o arquivo NFSe_Recebidas_*.xlsx mais recente."""
    candidatos = []
    for nome in os.listdir(pasta):
        low = nome.lower()
        if low.startswith("nfse") and low.endswith(".xlsx"):
            candidatos.append(os.path.join(pasta, nome))
    if not candidatos:
        return None
    return max(candidatos, key=os.path.getmtime)

def importar_nfse(conn, caminho):
    """Lê o xlsx 'NFSe_Recebidas_*.xlsx' (DANFSe nacional) e marca
    em_prefeitura=1 em nfse_consolidada.

    Colunas esperadas (cabeçalho linha 1):
      0  Geração       (data/hora — aceita dd/mm/yy ou dd/mm/yyyy)
      1  Emitida por   ('CNPJ - RAZÃO SOCIAL')
      2  Competência   ('MM/AAAA')
      3  Preço (R$)
      4  Situação      (pula 'NFS-e Cancelada')
      5  DANFE         (URL com chave de 50 dígitos no final)
    """
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb.active
    inseridos = atualizados = ignorados = canceladas = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        data_ger    = _data_iso(row[0])
        cnpj_emit, emitente = _parse_emitida_por(row[1])
        competencia = _s(row[2])
        valor       = _num_br(row[3])
        situacao    = _s(row[4])
        url         = _s(row[5])

        chave = _chave_nfse_url(url)
        if not chave:
            if cnpj_emit and competencia:
                chave = f"NFSE-{cnpj_emit}-{(competencia or '').replace('/','')}-{int((valor or 0)*100)}"
            else:
                ignorados += 1
                continue

        # pula canceladas e remove do banco se existir
        if _is_cancelada(situacao):
            conn.execute("DELETE FROM nfse_consolidada WHERE chave=?", (chave,))
            canceladas += 1
            continue

        ja_existe = conn.execute(
            "SELECT 1 FROM nfse_consolidada WHERE chave=?", (chave,)
        ).fetchone()

        conn.execute("""
            INSERT INTO nfse_consolidada
                (chave, cnpj_emitente, emitente, valor, competencia,
                 data_emissao, situacao, danfse_url,
                 em_prefeitura, ultima_importacao)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1, datetime('now','localtime'))
            ON CONFLICT(chave) DO UPDATE SET
                cnpj_emitente  = COALESCE(excluded.cnpj_emitente, nfse_consolidada.cnpj_emitente),
                emitente       = COALESCE(excluded.emitente, nfse_consolidada.emitente),
                valor          = COALESCE(excluded.valor, nfse_consolidada.valor),
                competencia    = COALESCE(excluded.competencia, nfse_consolidada.competencia),
                data_emissao   = COALESCE(excluded.data_emissao, nfse_consolidada.data_emissao),
                situacao       = COALESCE(excluded.situacao, nfse_consolidada.situacao),
                danfse_url     = COALESCE(excluded.danfse_url, nfse_consolidada.danfse_url),
                em_prefeitura  = 1,
                ultima_importacao = datetime('now','localtime')
        """, (chave, cnpj_emit, emitente, valor, competencia,
              data_ger, situacao, url))

        if ja_existe: atualizados += 1
        else:         inseridos += 1

    wb.close()
    return {"inseridos": inseridos, "atualizados": atualizados,
            "ignorados": ignorados, "canceladas": canceladas}

def _abrir_csv(caminho):
    """Abre o CSV tentando encodings comuns para texto pt-BR."""
    for enc in ("utf-8-sig", "cp1252", "latin-1", "utf-8"):
        try:
            f = open(caminho, "r", encoding=enc, newline="")
            f.readline()
            f.seek(0)
            return f, enc
        except UnicodeDecodeError:
            continue
    raise RuntimeError(f"Não consegui ler {caminho} com encodings comuns.")

def importar_sistema(conn, caminho):
    """Lê o CSV do sistema interno e marca em_sistema=1 nas notas encontradas.

    Cabeçalho (delimitador ';'):
      Ano;Mes;Dia;Data Entrada;Número Nota;Data Emissão;Usuário;Tipo Documento;
      Tipo Vencimento;Código Empresa;Nome Empresa;Razão Social;Inscr.Est.;
      Código Estabelecimento;Cidade;UF;Natureza;Descrição;CFOP;
      Valor Produtos;Valor Faturado;Valor Contábil;...;Chave
    """
    f, _enc = _abrir_csv(caminho)
    try:
        reader = csv.DictReader(f, delimiter=";")
        # mapeia nomes possíveis ignorando acentos/maiúsculas — só precisamos da Chave
        def col(d, *names):
            for n in names:
                for k in d.keys():
                    if k and k.strip().lower() == n.lower():
                        return d[k]
            return None

        ins_nfe = atu_nfe = ins_cte = atu_cte = ins_nfse = atu_nfse = ignorados = 0
        for linha in reader:
            chave = _chave(col(linha, "Chave"))
            numero       = _s(col(linha, "Número Nota", "Numero Nota"))
            data_emissao = _data_iso(col(linha, "Data Emissão", "Data Emissao"))
            valor        = _num_br(col(linha, "Valor Contábil", "Valor Contabil",
                                       "Valor Faturado", "Valor Produtos"))
            emitente     = _s(col(linha, "Razão Social", "Razao Social", "Nome Empresa"))
            usuario      = _s(col(linha, "Usuário", "Usuario"))

            if not chave:
                # CSV sem chave eletrônica → tenta MATCH em 3 lugares antes
                # de criar NFS sintética:
                #   1) NFe existente em nota_consolidada por nº + razão OU nº + valor
                #   2) NFSe na planilha da prefeitura por CNPJ/razão fuzzy
                #   3) fallback: cria NFS sintética em nota_consolidada
                cod_emp = _s(col(linha, "Código Empresa", "Codigo Empresa"))
                doc_csv = _doc_no_nome(emitente)
                norm_csv = _norm_emitente(emitente)

                # 1) NFe existente em nota_consolidada (por nº+razão ou nº+valor)
                if numero:
                    num_dig = re.sub(r"[^0-9]", "", numero)
                    if num_dig:
                        candidatos = conn.execute("""
                            SELECT chave, emitente, valor FROM nota_consolidada
                            WHERE numero = ? AND length(chave) = 44
                        """, (num_dig,)).fetchall()
                        chave_nfe_match = None
                        prefix = norm_csv[:8] if norm_csv and len(norm_csv) >= 6 else None
                        if prefix:
                            for c in candidatos:
                                if _norm_emitente(c[1] or "").startswith(prefix):
                                    chave_nfe_match = c[0]; break
                        if not chave_nfe_match and valor is not None:
                            for c in candidatos:
                                if c[2] is not None and abs(float(valor) - float(c[2])) < 0.01:
                                    chave_nfe_match = c[0]; break
                        if chave_nfe_match:
                            conn.execute("""
                                UPDATE nota_consolidada SET
                                    em_sistema = 1,
                                    usuario_lancamento = COALESCE(?, usuario_lancamento),
                                    ultima_importacao = datetime('now','localtime')
                                WHERE chave = ?
                            """, (usuario, chave_nfe_match))
                            atu_nfe += 1
                            continue

                # 2) NFSe da planilha da prefeitura
                chave_nfse = None
                if doc_csv:
                    r = conn.execute("""
                        SELECT chave FROM nfse_consolidada
                        WHERE em_prefeitura=1 AND (
                              cnpj_emitente = ?
                           OR cnpj_emitente LIKE ? || '%'
                           OR emitente LIKE '%' || ? || '%'
                        ) LIMIT 1
                    """, (doc_csv, doc_csv, doc_csv)).fetchone()
                    if r: chave_nfse = r[0]
                if not chave_nfse and norm_csv and len(norm_csv) >= 6:
                    prefix = norm_csv[:8]
                    for r in conn.execute("SELECT chave, emitente FROM nfse_consolidada WHERE em_prefeitura=1"):
                        if _norm_emitente(r[1] or "").startswith(prefix):
                            chave_nfse = r[0]; break

                if chave_nfse:
                    conn.execute("""
                        UPDATE nfse_consolidada SET
                            em_sistema = 1,
                            usuario_lancamento = COALESCE(?, usuario_lancamento),
                            ultima_importacao = datetime('now','localtime')
                        WHERE chave = ?
                    """, (usuario, chave_nfse))
                    atu_nfse += 1
                    continue

                # 3) fallback sintético em nota_consolidada
                if cod_emp and numero:
                    chave = f"NFS-{re.sub(r'[^A-Za-z0-9]', '', cod_emp)}-{re.sub(r'[^0-9]', '', numero)}"
                    modelo = "NFS"
                else:
                    ignorados += 1
                    continue
            else:
                modelo = chave[20:22]   # 55=NFe, 57=CTe

            if modelo == "57":
                ja_existe = conn.execute(
                    "SELECT 1 FROM cte_consolidada WHERE chave=?", (chave,)
                ).fetchone()
                conn.execute("""
                    INSERT INTO cte_consolidada
                        (chave, numero, emitente, valor, data_emissao,
                         em_sistema, usuario_lancamento, ultima_importacao)
                    VALUES (?, ?, ?, ?, ?, 1, ?, datetime('now','localtime'))
                    ON CONFLICT(chave) DO UPDATE SET
                        numero             = COALESCE(cte_consolidada.numero, excluded.numero),
                        emitente           = COALESCE(cte_consolidada.emitente, excluded.emitente),
                        valor              = COALESCE(cte_consolidada.valor, excluded.valor),
                        data_emissao       = COALESCE(cte_consolidada.data_emissao, excluded.data_emissao),
                        em_sistema         = 1,
                        usuario_lancamento = COALESCE(excluded.usuario_lancamento, cte_consolidada.usuario_lancamento),
                        ultima_importacao  = datetime('now','localtime')
                """, (chave, numero, emitente, valor, data_emissao, usuario))
                if ja_existe: atu_cte += 1
                else:         ins_cte += 1
            else:
                ja_existe = conn.execute(
                    "SELECT 1 FROM nota_consolidada WHERE chave=?", (chave,)
                ).fetchone()
                conn.execute("""
                    INSERT INTO nota_consolidada
                        (chave, numero, emitente, valor, data_emissao,
                         em_sistema, usuario_lancamento, ultima_importacao)
                    VALUES (?, ?, ?, ?, ?, 1, ?, datetime('now','localtime'))
                    ON CONFLICT(chave) DO UPDATE SET
                        numero             = COALESCE(nota_consolidada.numero, excluded.numero),
                        emitente           = COALESCE(nota_consolidada.emitente, excluded.emitente),
                        valor              = COALESCE(nota_consolidada.valor, excluded.valor),
                        data_emissao       = COALESCE(nota_consolidada.data_emissao, excluded.data_emissao),
                        em_sistema         = 1,
                        usuario_lancamento = COALESCE(excluded.usuario_lancamento, nota_consolidada.usuario_lancamento),
                        ultima_importacao  = datetime('now','localtime')
                """, (chave, numero, emitente, valor, data_emissao, usuario))
                if ja_existe: atu_nfe += 1
                else:         ins_nfe += 1
    finally:
        f.close()
    return {
        "nfe":  {"inseridos": ins_nfe,  "atualizados": atu_nfe},
        "cte":  {"inseridos": ins_cte,  "atualizados": atu_cte},
        "nfse": {"inseridos": ins_nfse, "atualizados": atu_nfse},
        "ignorados": ignorados,
    }

# ── orquestração ─────────────────────────────────────────────────────────────

def executar_importacao(caminho_sefaz, caminho_sistema, db_path=DB_PATH,
                        caminho_cte=None, caminho_nfse=None):
    """Importa até 4 fontes (NFe SEFAZ, CTe SEFAZ, NFSe prefeitura, CSV ERP)
    e registra a execução.

    Comportamento ACUMULATIVO: importar uma planilha NÃO zera as flags. Cada
    parâmetro é opcional. Notas canceladas (status 'Cancelada' / 'NFS-e
    Cancelada') são puladas e removidas do banco se já estavam lá.

    Linhas sem chave no CSV (NF-S) tentam casar com:
      (a) NFe existente em nota_consolidada por nº+razão ou nº+valor
      (b) NFSe da prefeitura por CNPJ/razão fuzzy
      (c) fallback: NFS sintética em nota_consolidada

    IMPORTANTE: importar a planilha NFSe ANTES do CSV maximiza o match.
    """
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    criar_banco(conn)

    r_sefaz = {"inseridos": 0, "atualizados": 0, "ignorados": 0}
    r_cte   = {"inseridos": 0, "atualizados": 0, "ignorados": 0}
    r_nfse  = {"inseridos": 0, "atualizados": 0, "ignorados": 0}
    r_sis   = {"nfe":  {"inseridos": 0, "atualizados": 0},
               "cte":  {"inseridos": 0, "atualizados": 0},
               "nfse": {"inseridos": 0, "atualizados": 0},
               "ignorados": 0}

    if caminho_sefaz and os.path.exists(caminho_sefaz):
        print(f"→ NFe SEFAZ: {caminho_sefaz}")
        r_sefaz = importar_sefaz(conn, caminho_sefaz)
        print(f"  inseridos={r_sefaz['inseridos']}  atualizados={r_sefaz['atualizados']}  "
              f"canceladas={r_sefaz.get('canceladas',0)}  ignorados={r_sefaz['ignorados']}")

    if caminho_cte and os.path.exists(caminho_cte):
        print(f"→ CT-e SEFAZ: {caminho_cte}")
        r_cte = importar_cte(conn, caminho_cte)
        print(f"  inseridos={r_cte['inseridos']}  atualizados={r_cte['atualizados']}  "
              f"canceladas={r_cte.get('canceladas',0)}  ignorados={r_cte['ignorados']}")

    if caminho_nfse and os.path.exists(caminho_nfse):
        print(f"→ NFS-e Prefeitura: {caminho_nfse}")
        r_nfse = importar_nfse(conn, caminho_nfse)
        print(f"  inseridos={r_nfse['inseridos']}  atualizados={r_nfse['atualizados']}  "
              f"canceladas={r_nfse.get('canceladas',0)}  ignorados={r_nfse['ignorados']}")
        movidas = migrar_nfs_sinteticas(conn)
        if movidas:
            print(f"  migração: {movidas} NFS sintéticas movidas pra nfse_consolidada")

    if caminho_sistema and os.path.exists(caminho_sistema):
        print(f"→ Sistema (ERP): {caminho_sistema}")
        r_sis = importar_sistema(conn, caminho_sistema)
        print(f"  NFe: ins={r_sis['nfe']['inseridos']} atu={r_sis['nfe']['atualizados']}"
              f" | CTe: ins={r_sis['cte']['inseridos']} atu={r_sis['cte']['atualizados']}"
              f" | NFSe match: atu={r_sis['nfse']['atualizados']}"
              f" | ignorados={r_sis['ignorados']}")
        consol = consolidar_nfs_com_nfes(conn)
        if consol:
            print(f"  consolidação: {consol} NFS sintéticas duplicadas removidas (eram NFes reais)")

    notas_sefaz   = conn.execute("SELECT COUNT(*) FROM nota_consolidada WHERE em_sefaz=1").fetchone()[0]
    notas_sistema = conn.execute("SELECT COUNT(*) FROM nota_consolidada WHERE em_sistema=1").fetchone()[0]
    notas_total   = conn.execute("SELECT COUNT(*) FROM nota_consolidada").fetchone()[0]
    ctes_sefaz    = conn.execute("SELECT COUNT(*) FROM cte_consolidada WHERE em_sefaz=1").fetchone()[0]
    ctes_sistema  = conn.execute("SELECT COUNT(*) FROM cte_consolidada WHERE em_sistema=1").fetchone()[0]
    ctes_total    = conn.execute("SELECT COUNT(*) FROM cte_consolidada").fetchone()[0]
    nfses_pref    = conn.execute("SELECT COUNT(*) FROM nfse_consolidada WHERE em_prefeitura=1").fetchone()[0]
    nfses_sis     = conn.execute("SELECT COUNT(*) FROM nfse_consolidada WHERE em_sistema=1").fetchone()[0]
    nfses_total   = conn.execute("SELECT COUNT(*) FROM nfse_consolidada").fetchone()[0]

    conn.execute("""
        INSERT INTO importacao (arquivo_sefaz, arquivo_cte, arquivo_nfse, arquivo_sistema,
                                notas_sefaz, ctes_sefaz, nfses_prefeitura,
                                notas_sistema, ctes_sistema, nfses_sistema,
                                notas_total, ctes_total, nfses_total)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (caminho_sefaz, caminho_cte, caminho_nfse, caminho_sistema,
          notas_sefaz, ctes_sefaz, nfses_pref,
          notas_sistema, ctes_sistema, nfses_sis,
          notas_total, ctes_total, nfses_total))

    conn.commit()
    conn.close()

    return {
        "sefaz":   r_sefaz,
        "cte":     r_cte,
        "nfse":    r_nfse,
        "sistema": r_sis,
        "totais": {
            "nfe":  {"em_sefaz": notas_sefaz, "em_sistema": notas_sistema, "total": notas_total},
            "cte":  {"em_sefaz": ctes_sefaz,  "em_sistema": ctes_sistema,  "total": ctes_total},
            "nfse": {"em_prefeitura": nfses_pref, "em_sistema": nfses_sis, "total": nfses_total},
        },
    }

def main():
    parser = argparse.ArgumentParser(description="Importa planilhas SEFAZ (NFe + CTe), NFS-e da prefeitura e CSV do ERP.")
    parser.add_argument("--sefaz",   help="Caminho do .xlsx FSist-NFe-Recebidas (auto-detecta na pasta se omitido)")
    parser.add_argument("--cte",     help="Caminho do .xlsx FSist-CTe (auto-detecta na pasta se omitido)")
    parser.add_argument("--nfse",    help="Caminho do .xlsx NFSe_Recebidas (auto-detecta na pasta se omitido)")
    parser.add_argument("--sistema", help="Caminho do .csv Notas de Entrada", default=ARQUIVO_SISTEMA_PADRAO)
    parser.add_argument("--vazio",   action="store_true",
                        help="Cria apenas o banco vazio (sem importar nada); use a tela web depois")
    args = parser.parse_args()

    if args.vazio:
        conn = sqlite3.connect(DB_PATH)
        criar_banco(conn)
        conn.close()
        print(f"Banco vazio criado em {DB_PATH}. Rode  python app.py  e importe pela tela web.")
        return

    sefaz   = args.sefaz   or detectar_arquivo_sefaz(".")
    cte     = args.cte     or detectar_arquivo_cte(".")
    nfse    = args.nfse    or detectar_arquivo_nfse(".")
    sistema = args.sistema if (args.sistema and os.path.exists(args.sistema)) else None

    if not sefaz and not cte and not nfse and not sistema:
        print(f"Nenhuma planilha encontrada na pasta atual.\n"
              f"  - NFe SEFAZ:   FSist-NFe-Recebidas-*.xlsx (ou --sefaz)\n"
              f"  - CT-e SEFAZ:  FSist-CTe-*.xlsx (ou --cte)\n"
              f"  - NFS-e:       NFSe_Recebidas_*.xlsx (ou --nfse)\n"
              f"  - ERP:         '{ARQUIVO_SISTEMA_PADRAO}' (ou --sistema)\n"
              f"Para criar apenas o banco vazio e importar pela tela web, rode:\n"
              f"  python init_db.py --vazio", file=sys.stderr)
        sys.exit(1)

    resumo = executar_importacao(sefaz, sistema, caminho_cte=cte, caminho_nfse=nfse)

    print("\n── Resumo ──────────────────────────────────────────────")
    print(f"  NFe  — em SEFAZ:      {resumo['totais']['nfe']['em_sefaz']:>6}  em Sistema: {resumo['totais']['nfe']['em_sistema']:>6}  total: {resumo['totais']['nfe']['total']:>6}")
    print(f"  CTe  — em SEFAZ:      {resumo['totais']['cte']['em_sefaz']:>6}  em Sistema: {resumo['totais']['cte']['em_sistema']:>6}  total: {resumo['totais']['cte']['total']:>6}")
    print(f"  NFSe — em Prefeitura: {resumo['totais']['nfse']['em_prefeitura']:>6}  em Sistema: {resumo['totais']['nfse']['em_sistema']:>6}  total: {resumo['totais']['nfse']['total']:>6}")
    print("Pronto. Rode  python app.py  e acesse http://localhost:5001")

if __name__ == "__main__":
    main()
